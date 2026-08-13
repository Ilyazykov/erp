import csv, re, time
from datetime import date, timedelta
from calendar import monthrange

SCRATCHPAD = "/private/tmp/claude-501/-Users-ilyazykov-code-personal-buildit/ffcba325-4fdb-4f82-b4b4-957caab60f07/scratchpad"

WEIGHTS = {
    'SBER': 0.4172,
    'YDEX': 0.3152,
    'T':    0.1744,
    'OZON': 0.0759,
    'ROSN': 0.0099,
    'VTBR': 0.0074,
}

SHARES = {
    'SBER': 21586948000,
    'ROSN': 10598177817,
    'VTBR': 12927766416,
    'YDEX': 396012957,
    'T':    2682747860,
    'OZON': 208992107,
}

# YNDX (old ticker) had 326M shares; YDEX (post-redomicile) has 396M
YNDX_SHARES = 326342270
YDEX_START = date(2024, 7, 24)  # first trading day of YDEX on MOEX

# TCSG/T share history:
# - Until Nov 27 2024: TCSG, 199.3M shares
# - Nov 28 2024 – Apr 16 2026: T after redomicile + Rosbank merger, ~268.3M shares
# - From Apr 17 2026: T after 1:9.58 split, 2682.7M shares (current ISSUESIZE)
TCSG_SHARES = 199305492
T_START      = date(2024, 11, 28)
T_SPLIT_DATE = date(2026, 4, 17)
T_PRE_SPLIT_SHARES = 268274786   # = 2682747860 / 9.58 (implied from price ratio)

# VTBR did a 1:4664 reverse split on 2024-07-15.
# Before that date: shares were ~60.3 trillion (= 12.9B × 4664)
# After: 12.9B shares at ~93 rub each.
# We handle this by using adjusted shares count per period.
VTBR_SPLIT_DATE = date(2024, 7, 15)
VTBR_SPLIT_RATIO = 4664  # old shares per 1 new share

# Quarter end dates
QUARTER_END = {1: (3, 31), 2: (6, 30), 3: (9, 30), 4: (12, 31)}

def quarter_str_to_report_date(qstr):
    """
    'SBER' публикует примерно:
    Q1 → конец апреля, Q2 → конец июля, Q3 → конец октября, Q4 → конец февраля следующего года
    Используем консервативную оценку: 45 дней после конца квартала.
    Исключения для 2022 (Сбер/ВТБ не публиковали) — обрабатываем через None значения.
    """
    year = int(qstr[:4])
    q = int(qstr[5])
    end_month, end_day = QUARTER_END[q]
    if q == 4:
        # Q4: годовые данные известны с 1 января следующего года
        report_date = date(year + 1, 1, 1)
    else:
        qend = date(year, end_month, end_day)
        report_date = qend + timedelta(days=45)
    return report_date

def load_quarterly_ni():
    """Load from CSV, convert to dict: ticker → list of (qstr, value, report_date)"""
    data = {}
    with open(f"{SCRATCHPAD}/quarterly_ni.csv") as f:
        for row in csv.DictReader(f):
            ticker = row['ticker']
            qstr = row['quarter']
            v = row['net_income_bln']
            ni = float(v) if v else None
            rdate = quarter_str_to_report_date(qstr)
            if ticker not in data:
                data[ticker] = []
            data[ticker].append((qstr, ni, rdate))
    return data

def get_ttm_ni(quarterly_data, ticker, as_of):
    """TTM using only quarters published by as_of date. Returns None if <4 quarters."""
    quarters = quarterly_data.get(ticker, [])
    available = [(q, ni, rd) for q, ni, rd in quarters
                 if rd <= as_of and ni is not None]
    if len(available) < 4:
        return None
    available.sort(key=lambda x: x[0])
    last4 = available[-4:]
    return sum(x[1] for x in last4)

def load_prices(ticker):
    prices = {}
    with open(f"{SCRATCHPAD}/prices_{ticker}.csv") as f:
        for row in csv.DictReader(f):
            prices[row['date']] = float(row['close'])
    # For YDEX: also load YNDX historical prices (pre-redomicile)
    if ticker == 'YDEX':
        try:
            with open(f"{SCRATCHPAD}/prices_YNDX.csv") as f:
                for row in csv.DictReader(f):
                    d = row['date']
                    if d not in prices:
                        prices[d] = float(row['close'])
        except FileNotFoundError:
            pass
    if ticker == 'T':
        try:
            with open(f"{SCRATCHPAD}/prices_TCSG.csv") as f:
                for row in csv.DictReader(f):
                    d = row['date']
                    if d not in prices:
                        prices[d] = float(row['close'])
        except FileNotFoundError:
            pass
    return prices

def last_price_of_month(prices, year, month):
    days_in_month = monthrange(year, month)[1]
    for day in range(days_in_month, 0, -1):
        ds = date(year, month, day).isoformat()
        if ds in prices:
            return prices[ds]
    return None

def load_ofz():
    ofz = {}
    with open(f"{SCRATCHPAD}/ofz10y_monthly.csv") as f:
        for row in csv.DictReader(f):
            if row['ofz10y_pct']:
                ofz[row['date']] = float(row['ofz10y_pct'])
    return ofz

def main():
    quarterly_ni = load_quarterly_ni()
    all_prices = {t: load_prices(t) for t in WEIGHTS}
    ofz_data = load_ofz()

    # Monthly range: 2019-01 to 2026-07
    months = []
    y, m = 2019, 1
    while (y, m) <= (2026, 7):
        months.append((y, m))
        m += 1
        if m > 12: m, y = 1, y + 1

    results = []
    for (y, m) in months:
        as_of = date(y, m, monthrange(y, m)[1])
        month_str = f"{y}-{m:02d}"

        # OFZ
        ofz_val = None
        for day in range(1, 10):
            try:
                k = date(y, m, day).isoformat()
                if k in ofz_data:
                    ofz_val = ofz_data[k]
                    break
            except:
                pass

        ey_values = {}
        for ticker in WEIGHTS:
            price = last_price_of_month(all_prices[ticker], y, m)
            ttm_ni = get_ttm_ni(quarterly_ni, ticker, as_of)
            if price is None or ttm_ni is None:
                ey_values[ticker] = None
                continue
            shares = SHARES[ticker]
            if ticker == 'VTBR' and as_of < VTBR_SPLIT_DATE:
                shares = shares * VTBR_SPLIT_RATIO
            if ticker == 'YDEX' and as_of < YDEX_START:
                shares = YNDX_SHARES
            if ticker == 'T' and as_of < T_START:
                shares = TCSG_SHARES
            elif ticker == 'T' and as_of < T_SPLIT_DATE:
                shares = T_PRE_SPLIT_SHARES
            mcap = price * shares / 1e9  # млрд руб
            ey_values[ticker] = ttm_ni / mcap * 100  # %

        available = [t for t in WEIGHTS if ey_values.get(t) is not None]
        if not available:
            port_ey = None
        else:
            total_w = sum(WEIGHTS[t] for t in available)
            port_ey = sum(WEIGHTS[t] / total_w * ey_values[t] for t in available)

        erp = (port_ey - ofz_val) if (port_ey is not None and ofz_val is not None and ofz_val > 0) else None

        results.append({
            'date': month_str,
            'SBER_ey': ey_values.get('SBER'),
            'YDEX_ey': ey_values.get('YDEX'),
            'T_ey':    ey_values.get('T'),
            'OZON_ey': ey_values.get('OZON'),
            'ROSN_ey': ey_values.get('ROSN'),
            'VTBR_ey': ey_values.get('VTBR'),
            'portfolio_ey': port_ey,
            'ofz10y':  ofz_val,
            'erp':     erp,
        })

    return results

rows = main()

# ── Print first 20 non-empty rows ─────────────────────────────────────────────
def fmt(v):
    return f"{v:7.2f}" if v is not None else "   None"

print(f"{'Date':<8} {'SBER':>7} {'YDEX':>7} {'T':>7} {'OZON':>7} "
      f"{'ROSN':>7} {'VTBR':>7} {'PortEY':>7} {'OFZ10Y':>7} {'ERP':>7}")
print("-" * 80)
shown = 0
for r in rows:
    if r['portfolio_ey'] is not None:
        print(f"{r['date']:<8} {fmt(r['SBER_ey'])} {fmt(r['YDEX_ey'])} "
              f"{fmt(r['T_ey'])} {fmt(r['OZON_ey'])} {fmt(r['ROSN_ey'])} "
              f"{fmt(r['VTBR_ey'])} {fmt(r['portfolio_ey'])} "
              f"{fmt(r['ofz10y'])} {fmt(r['erp'])}")
        shown += 1
        if shown >= 20:
            break

print(f"\nTotal months: {len(rows)}, with data: {sum(1 for r in rows if r['portfolio_ey'] is not None)}")

# ── Save CSV ──────────────────────────────────────────────────────────────────
out_csv = "/Users/ilyazykov/Downloads/erp_portfolio.csv"
fields = ['date','SBER_ey','YDEX_ey','T_ey','OZON_ey','ROSN_ey','VTBR_ey',
          'portfolio_ey','ofz10y','erp']
with open(out_csv, 'w', newline='') as f:
    w = csv.DictWriter(f, fieldnames=fields)
    w.writeheader()
    for r in rows:
        w.writerow({k: (round(r[k], 4) if (r[k] is not None and k != 'date') else (r[k] or ''))
                    for k in fields})
print(f"CSV: {out_csv}")

# ── Save Excel ────────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ERP Portfolio"

    headers = ['Дата','SBER EY%','YDEX EY%','T EY%','OZON EY%',
               'ROSN EY%','VTBR EY%','Portfolio EY%','OFZ 10Y%','ERP Proxy%']
    hdr_fill = PatternFill("solid", fgColor="1e2a3a")
    hdr_font = Font(bold=True, color="e6edf3")

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = 13

    for ri, r in enumerate(rows, 2):
        vals = [r['date'], r['SBER_ey'], r['YDEX_ey'], r['T_ey'], r['OZON_ey'],
                r['ROSN_ey'], r['VTBR_ey'], r['portfolio_ey'], r['ofz10y'], r['erp']]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            if v is not None and ci > 1:
                c.number_format = '0.00'
                if ci == 10 and v is not None:  # ERP column
                    c.fill = PatternFill("solid", fgColor="c6efce" if v > 0 else "ffc7ce")

    wb.save("/Users/ilyazykov/Downloads/erp_portfolio.xlsx")
    print("Excel: /Users/ilyazykov/Downloads/erp_portfolio.xlsx")
except ImportError:
    print("openpyxl not available")

# ── Chart ─────────────────────────────────────────────────────────────────────
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from datetime import datetime

plot_rows = [r for r in rows if r['portfolio_ey'] is not None and r['ofz10y'] is not None]
dates_p = [datetime.strptime(r['date'], '%Y-%m') for r in plot_rows]
ey_p    = [r['portfolio_ey'] for r in plot_rows]
ofz_p   = [r['ofz10y'] for r in plot_rows]
erp_p   = [r['erp'] for r in plot_rows]

fig, ax = plt.subplots(figsize=(16, 7))
fig.patch.set_facecolor('#0d1117')
ax.set_facecolor('#0d1117')

ax.plot(dates_p, ey_p,  color='#4da8c8', linewidth=2,   label='Portfolio Earnings Yield')
ax.plot(dates_p, ofz_p, color='#f0b429', linewidth=2,   label='OFZ 10Y Yield')
ax.plot(dates_p, erp_p, color='#3dd68c', linewidth=2.5, label='ERP Proxy (EY − OFZ)')
ax.axhline(0, color='#6a7381', linewidth=1, linestyle='--', alpha=0.7)

ax.fill_between(dates_p, erp_p, 0,
    where=[e >= 0 for e in erp_p], alpha=0.12, color='#3dd68c', label='ERP > 0 (недооценка)')
ax.fill_between(dates_p, erp_p, 0,
    where=[e < 0 for e in erp_p],  alpha=0.12, color='#e05252', label='ERP < 0 (переоценка)')

ax.set_title('ERP Proxy российского портфеля (SBER/YDEX/T/OZON/ROSN/VTBR)',
             color='#e6edf3', fontsize=13, pad=12)
ax.set_ylabel('%, годовых', color='#c9d1d9', fontsize=10)
ax.tick_params(colors='#6a7381', labelsize=8)
for spine in ax.spines.values():
    spine.set_edgecolor('#1e2830')
ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
ax.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
plt.xticks(rotation=45, ha='right')
ax.legend(facecolor='#13191f', edgecolor='#1e2830', labelcolor='#c9d1d9', fontsize=9)
ax.grid(axis='y', color='#1e2830', linewidth=0.5, alpha=0.7)
ax.grid(axis='x', color='#1e2830', linewidth=0.3, alpha=0.5)

plt.tight_layout()
plt.savefig("/Users/ilyazykov/Downloads/erp_portfolio.png", dpi=150,
            bbox_inches='tight', facecolor='#0d1117')
print("Chart: /Users/ilyazykov/Downloads/erp_portfolio.png")
